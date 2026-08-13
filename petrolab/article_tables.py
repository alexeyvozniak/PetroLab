from __future__ import annotations

from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from petrolab.column_schema import describe_header
from petrolab.visualization_presets import TABLE_PRESETS, TablePreset


IDENTIFIER_COLUMNS = {
    "Project",
    "Проект",
    "Rock",
    "Sample",
    "Grain",
    "Point",
    "Generation",
    "Набор",
    "Минерал",
    "Massif",
    "Lithology",
}


def _is_trace(column: str) -> bool:
    """Return True only for a concentration column with an explicit trace-element unit.

    Do not infer trace chemistry from short letter substrings in arbitrary derived
    column names (for example ``AlIV`` contains ``Li`` case-insensitively). The same
    header semantics used by the importer are the source of truth here as well.
    """
    descriptor = describe_header(column)
    return descriptor.quantity_kind in {"trace_element", "element_concentration"}


def _is_identifier(column: object) -> bool:
    return str(column) in IDENTIFIER_COLUMNS or str(column).startswith("_")


def format_dataframe_for_article(
    dataframe: pd.DataFrame,
    *,
    preset_name: str,
    columns: Iterable[str] | None = None,
) -> pd.DataFrame:
    preset = TABLE_PRESETS[preset_name]
    result = dataframe[list(columns)].copy() if columns is not None else dataframe.copy()
    for column in result.columns:
        if _is_identifier(column):
            # Identifiers are labels, not measurements. Numeric-looking IDs such as
            # "001" or "007" must keep their exact textual representation.
            continue
        original = result[column].copy()
        values = pd.to_numeric(original, errors="coerce")
        numeric_mask = values.notna()
        if not numeric_mask.any():
            continue
        decimals = preset.decimals_trace if _is_trace(str(column)) else preset.decimals_major
        # Round only cells that are genuinely numeric. Qualifiers such as '<0.01',
        # 'bdl' or analytical comments are scientific information and must survive.
        rounded = values.round(decimals)
        if numeric_mask.all():
            result[column] = rounded
        else:
            mixed = original.astype(object)
            mixed.loc[numeric_mask] = rounded.loc[numeric_mask]
            result[column] = mixed
    return result


def article_table_xlsx_bytes(
    dataframe: pd.DataFrame,
    *,
    preset_name: str,
    sheet_name: str = "Table",
    title: str = "",
    note: str = "",
) -> bytes:
    preset: TablePreset = TABLE_PRESETS[preset_name]
    buffer = BytesIO()
    startrow = 2 if title else 0
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name[:31], startrow=startrow)
        ws = writer.book[sheet_name[:31]]
        thin = Side(style="thin", color="808080")
        if title:
            ws.cell(1, 1, title)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(dataframe.columns)))
            ws.cell(1, 1).font = Font(name=preset.font_family, size=preset.header_size + 1, bold=True)
        header_row = startrow + 1
        for cell in ws[header_row]:
            cell.font = Font(name=preset.font_family, size=preset.header_size, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
            cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = Font(name=preset.font_family, size=preset.font_size)
                cell.alignment = Alignment(vertical="center")
        for idx, column in enumerate(dataframe.columns, start=1):
            sample = [str(column)] + [str(value) for value in dataframe[column].head(100).dropna().tolist()]
            width = min(max(max((len(value) for value in sample), default=8) + 2, 8), 24)
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = ws.cell(header_row + 1, 1)
        if preset.repeat_header:
            ws.print_title_rows = f"{header_row}:{header_row}"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape" if preset.landscape else "portrait"
        if note:
            note_row = header_row + len(dataframe) + 2
            ws.cell(note_row, 1, note)
            ws.cell(note_row, 1).font = Font(name=preset.font_family, size=max(7, preset.font_size - 1), italic=True)
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=max(1, len(dataframe.columns)))
    return buffer.getvalue()

"""Reproducibility manifests shared by figures and supplementary tables."""
from __future__ import annotations

import json
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from petrolab.db import _utcnow
from petrolab.derived import formula_provenance_rows


def build_selection_manifest(
    *,
    kind: str,
    dataframe: pd.DataFrame,
    dataset_ids: list[int],
    filters: dict[str, Any] | None = None,
    recipe: dict[str, Any] | None = None,
    overlay_source: str = "",
) -> dict[str, Any]:
    """Describe exactly which stored observations produced an export."""
    ids = (
        dataframe["_analysis_id"].dropna().astype(str).tolist()
        if "_analysis_id" in dataframe.columns else []
    )
    sources = []
    for column in ("Проект", "Набор", "Источник", "Лист"):
        if column in dataframe.columns:
            sources.append({"field": column, "values": sorted(dataframe[column].dropna().astype(str).unique().tolist())})
    return {
        "schema": "petrolab-publication-manifest/v1",
        "created_at": _utcnow(),
        "kind": str(kind),
        "analysis_ids": ids,
        "dataset_ids": sorted({int(value) for value in dataset_ids}),
        "filters": filters or {},
        "recipe": recipe or {},
        "formula_provenance": formula_provenance_rows(dataset_ids),
        "overlay_source": str(overlay_source),
        "sources": sources,
    }


def manifest_json_bytes(manifest: dict[str, Any]) -> bytes:
    return json.dumps(manifest, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def workbook_with_manifest(
    sheets: dict[str, pd.DataFrame],
    manifest: dict[str, Any],
) -> bytes:
    """Export data and a machine-readable manifest in the same XLSX file."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, dataframe in sheets.items():
            dataframe.to_excel(writer, index=False, sheet_name=str(name)[:31])
        pd.DataFrame([{"manifest_json": manifest_json_bytes(manifest).decode("utf-8")}]).to_excel(
            writer, index=False, sheet_name="Manifest"
        )
    return buffer.getvalue()


def append_manifest_to_xlsx(workbook_bytes: bytes, manifest: dict[str, Any]) -> bytes:
    """Add provenance without replacing the journal formatting of an XLSX table."""
    book = load_workbook(BytesIO(workbook_bytes))
    if "Manifest" in book.sheetnames:
        del book["Manifest"]
    sheet = book.create_sheet("Manifest")
    sheet.append(["manifest_json"])
    sheet.append([manifest_json_bytes(manifest).decode("utf-8")])
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()

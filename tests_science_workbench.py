from __future__ import annotations

import io
import os
import sqlite3
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="petrolab_science_") as tmp:
        root = Path(tmp)
        os.environ["PETROLAB_DATA_DIR"] = str(root / "data")

        from petrolab.article_tables import article_table_xlsx_bytes, format_dataframe_for_article
        from petrolab.db import DB_PATH, create_project, ensure_storage
        from petrolab.extended_plotting import (
            CI_CHONDRITE_1995,
            PRIMITIVE_MANTLE_1989,
            build_histogram_figure,
            build_pattern_figure,
            prepare_pattern,
        )
        from petrolab.repositories.rock_repository import (
            composition_wide,
            create_rock,
            get_composition,
            get_isotopes,
            list_rocks,
            replace_composition,
            replace_isotopes,
        )
        from petrolab.rock_plotting import build_tas_figure, figure_bytes
        from petrolab.scientific_overlays import (
            FE2O3_TO_FEO_EQUIVALENT,
            XY_OVERLAYS,
            classify_grutter_g10,
            total_fe_as_feo,
        )
        from petrolab.scientific_plotting import build_scientific_xy_figure
        from petrolab.services.rock_image_service import list_rock_images, save_rock_image
        from petrolab.services.rock_service import (
            delete_rock_with_assets,
            import_rocks_wide,
            inferred_whole_rock_fe3_fraction,
            rhodes_equilibrium_fo,
            whole_rock_mg_number,
        )
        from petrolab.statistics import prepare_matrix, run_clustering, run_pca
        from petrolab.ui.pages.science_plots import _axis_candidates
        from petrolab.visualization_presets import (
            FIGURE_PRESETS,
            POINT_STYLE_PRESETS,
            SCIENTIFIC_PLOT_PRESETS,
            TABLE_PRESETS,
        )

        ensure_storage()
        con = sqlite3.connect(DB_PATH)
        try:
            tables = {row[0] for row in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        finally:
            con.close()
        assert {"rock_samples", "rock_compositions", "rock_isotopes", "rock_mineral_links", "rock_images"}.issubset(tables)

        project_id = create_project("Science test", "")
        rock_id = create_rock(project_id, "R1", massif="Kola", lithology="lamprophyre", age_ma=380.0)
        composition = {
            "SiO2": 44.0, "Na2O": 2.5, "K2O": 3.0, "MgO": 12.0, "FeOt": 10.0,
            "La [µg/g]": 23.7, "Ce [µg/g]": 61.3,
        }
        replace_composition(rock_id, composition, units={"La [µg/g]": "µg/g", "Ce [µg/g]": "µg/g"})
        stored = get_composition(rock_id)
        assert set(["SiO2", "MgO", "FeOt"]).issubset(set(stored["analyte"]))
        wide = composition_wide(project_id)
        assert len(wide) == 1 and float(wide.loc[0, "SiO2"]) == 44.0
        mgnum = whole_rock_mg_number(composition)
        assert 0.6 < mgnum < 0.8
        assert 80.0 < rhodes_equilibrium_fo(mgnum, 0.30) < 100.0

        split_rock_fe = {"MgO": 10.0, "FeO": 5.0, "Fe2O3": 2.0}
        inferred_fraction = inferred_whole_rock_fe3_fraction(split_rock_fe)
        assert inferred_fraction is not None and 0.0 < inferred_fraction < 1.0
        ferric_moles = 2.0 * 2.0 / 159.688
        ferrous_moles = 5.0 / 71.844
        expected_fraction = ferric_moles / (ferric_moles + ferrous_moles)
        assert np.isclose(inferred_fraction, expected_fraction)
        expected_mg = (10.0 / 40.304) / ((10.0 / 40.304) + ferrous_moles)
        assert np.isclose(whole_rock_mg_number(split_rock_fe), expected_mg)

        isotopes = pd.DataFrame([
            {"system": "Sr", "ratio_name": "87Sr/86Sr", "value": 0.7032, "uncertainty": 0.00002,
             "initial_value": 0.7029, "age_ma_used": 380.0, "method": "TIMS", "laboratory": "Lab", "notes": ""}
        ])
        replace_isotopes(rock_id, isotopes)
        assert len(get_isotopes(rock_id)) == 1

        incoming = pd.DataFrame({"Rock": ["R1", "R2"], "SiO2": [50.0, 48.0], "MgO": [9.0, 11.0]})
        skipped = import_rocks_wide(incoming, project_id=project_id, name_column="Rock", on_conflict="skip")
        assert skipped.skipped_names == ("R1",)
        assert len(skipped.created_ids) == 1
        r2_id = skipped.created_ids[0]
        updated = import_rocks_wide(
            pd.DataFrame({"Rock": ["R1"], "SiO2": [52.0], "MgO": [8.0]}),
            project_id=project_id,
            name_column="Rock",
            on_conflict="update",
        )
        assert updated.updated_ids == (rock_id,)
        r1_comp = get_composition(rock_id).set_index("analyte")
        assert float(r1_comp.loc["SiO2", "value"]) == 52.0
        before_duplicate = len(list_rocks(project_id))
        try:
            import_rocks_wide(
                pd.DataFrame({"Rock": ["R3", "R3"], "SiO2": [45.0, 46.0]}),
                project_id=project_id,
                name_column="Rock",
            )
        except ValueError:
            pass
        else:
            raise AssertionError("duplicate names inside one whole-rock import must fail preflight")
        assert len(list_rocks(project_id)) == before_duplicate

        image_buffer = io.BytesIO()
        Image.new("RGB", (8, 8), "white").save(image_buffer, format="PNG")
        save_rock_image(r2_id, "rock.png", image_buffer.getvalue())
        rock_assets = list_rock_images(r2_id)
        assert len(rock_assets) == 1
        image_path = Path(rock_assets[0]["stored_path"])
        assert image_path.exists()
        delete_rock_with_assets(r2_id)
        assert not image_path.exists()
        assert all(int(rock["id"]) != r2_id for rock in list_rocks(project_id))

        wide = composition_wide(project_id)
        fig = build_tas_figure(wide)
        assert len(figure_bytes(fig, "png", 150)) > 1000
        mono_tas = build_tas_figure(
            pd.DataFrame({
                "SiO2": [45.0, 50.0], "Na2O": [2.0, 3.0], "K2O": [1.0, 2.0],
                "Massif": ["A", "B"], "Rock": ["a", "b"],
            }),
            group_column="Massif",
            label_column=None,
            monochrome=True,
            point_style_name="balanced",
        )
        assert len(mono_tas.axes[0].collections) == 2
        for collection in mono_tas.axes[0].collections:
            edges = collection.get_edgecolors()
            assert len(edges) and np.allclose(edges[0][:3], [0.0, 0.0, 0.0])

        traces = pd.DataFrame({"La [µg/g]": [23.7, 47.4], "Ce [µg/g]": [61.3, 122.6], "Pr": [9.28, 18.56]})
        normalized = prepare_pattern(traces, ["La", "Ce", "Pr"], CI_CHONDRITE_1995)
        assert normalized.elements == ("La", "Ce")
        assert "Pr" in normalized.missing_elements, "Unknown-unit bare trace element must not be normalized"
        assert np.allclose(normalized.data["La"].to_numpy(), [100.0, 200.0])
        raw = prepare_pattern(traces, ["Pr"], None)
        assert raw.elements == ("Pr",)

        oxide_pattern = prepare_pattern(
            pd.DataFrame({"K2O": [1.0], "P2O5": [0.5], "TiO2": [2.0]}),
            ["K", "P", "Ti"],
            PRIMITIVE_MANTLE_1989,
        )
        assert oxide_pattern.elements == ("K", "P", "Ti")
        assert oxide_pattern.source_columns["K"].startswith("K2O wt.%")
        assert 30.0 < float(oxide_pattern.data.loc[0, "K"]) < 35.0
        assert 20.0 < float(oxide_pattern.data.loc[0, "P"]) < 25.0
        assert 8.0 < float(oxide_pattern.data.loc[0, "Ti"]) < 10.5
        mono_pattern = build_pattern_figure(normalized, monochrome=True)
        assert mono_pattern.axes[0].lines
        for line in mono_pattern.axes[0].lines:
            assert line.get_color() == "black"
        mono_hist = build_histogram_figure(
            pd.DataFrame({"X": [1.0, 2.0, 3.0, 4.0], "G": ["A", "A", "B", "B"]}),
            "X",
            group_column="G",
            monochrome=True,
        )
        assert mono_hist.axes[0].patches
        for patch in mono_hist.axes[0].patches:
            face = patch.get_facecolor()
            assert np.isclose(face[0], face[1]) and np.isclose(face[1], face[2])

        features = pd.DataFrame({"A": [1.0, 1.1, 5.0, 5.1], "B": [2.0, 2.1, 8.0, 8.1]})
        prepared = prepare_matrix(features, ["A", "B"], scaler="standard")
        pca = run_pca(prepared, 2)
        assert pca.scores.shape == (4, 2)
        clusters = run_clustering(prepared, method="kmeans", n_clusters=2)
        assert clusters.labels.nunique() == 2
        single = prepare_matrix(pd.DataFrame({"A": [1.0], "B": [2.0]}), ["A", "B"])
        for function in (run_pca, run_clustering):
            try:
                function(single)
            except ValueError:
                pass
            else:
                raise AssertionError("undersized statistical model must fail with a clear ValueError")

        table = format_dataframe_for_article(
            pd.DataFrame({
                "SiO2": [40.1234, 41.9876],
                "Rb [µg/g]": [123.456, "<0.01"],
                "AlIV": [1.2345, 1.9876],
            }),
            preset_name="Lithos",
        )
        assert float(table.loc[0, "SiO2"]) == 40.12
        assert float(table.loc[0, "Rb [µg/g]"]) == 123.5
        assert table.loc[1, "Rb [µg/g]"] == "<0.01", "Detection-limit qualifier must survive formatting"
        assert float(table.loc[0, "AlIV"]) == 1.23, "AlIV must not be mistaken for Li concentration"
        payload = article_table_xlsx_bytes(table, preset_name="Lithos", title="Test")
        path = root / "table.xlsx"
        path.write_bytes(payload)
        workbook = load_workbook(path)
        assert "Table" in workbook.sheetnames
        workbook.close()

        axis_frame = pd.DataFrame({
            "FeO": [8.0, 9.0], "NiO": [0.2, 0.3], "Ni [µg/g]": [1800.0, 2100.0],
        })
        numeric = list(axis_frame.columns)
        assert _axis_candidates(axis_frame, "FeOt", numeric) == []
        assert _axis_candidates(axis_frame, "Ni", numeric) == ["Ni [µg/g]"]

        assert {"Lithos", "ДАН", "Elsevier 1-column"}.issubset(FIGURE_PRESETS)
        assert {"balanced", "open", "bw"}.issubset(POINT_STYLE_PRESETS)
        assert "Lithos" in TABLE_PRESETS
        assert "olivine_fo_nio" in SCIENTIFIC_PLOT_PRESETS
        assert SCIENTIFIC_PLOT_PRESETS["olivine_fo_nio"].y == "NiO"
        assert SCIENTIFIC_PLOT_PRESETS["olivine_fo_ni"].y == "Ni"
        assert {"ilmenite_wyatt_kimberlite_curve", "garnet_grutter_g10_diagnostic"}.issubset(XY_OVERLAYS)

        garnet = pd.DataFrame({"CaO": [3.0], "Cr2O3": [6.0], "MnO": [0.2], "MgO": [20.0], "FeO": [8.0]})
        classification = classify_grutter_g10(garnet)
        assert classification.iloc[0] == "G10A diagnostic"
        split_fe = pd.DataFrame({"FeO": [6.0], "Fe2O3": [2.0]})
        expected_total = 6.0 + 2.0 * FE2O3_TO_FEO_EQUIVALENT
        assert np.isclose(float(total_fe_as_feo(split_fe).iloc[0]), expected_total)
        assert np.isclose(
            float(total_fe_as_feo(pd.DataFrame({"FeOt": [expected_total]})).iloc[0]),
            expected_total,
        )

        mono_fig = build_scientific_xy_figure(
            pd.DataFrame({"X": [1.0, 2.0], "Y": [2.0, 3.0], "Group": ["A", "B"]}),
            x="X", y="Y", x_label="X", y_label="Y", group_column="Group",
            point_style_name="balanced", monochrome=True,
        )
        assert len(mono_fig.axes[0].collections) == 2
        for collection in mono_fig.axes[0].collections:
            edges = collection.get_edgecolors()
            assert len(edges) and np.allclose(edges[0][:3], [0.0, 0.0, 0.0])

        for figure in [fig, mono_tas, mono_pattern, mono_hist, mono_fig]:
            plt = __import__("matplotlib.pyplot", fromlist=["close"])
            plt.close(figure)

    print("science workbench tests: OK")


if __name__ == "__main__":
    main()

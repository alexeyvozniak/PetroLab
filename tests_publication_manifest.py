from __future__ import annotations

import os
import tempfile
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="petrolab_manifest_", ignore_cleanup_errors=True) as tmp:
        os.environ["PETROLAB_DATA_DIR"] = str(Path(tmp) / "data")
        from petrolab.db import add_dataset, create_project, replace_dataset_rows
        from petrolab.publication_manifest import build_selection_manifest, workbook_with_manifest
        from petrolab.source_registry import (
            SOURCE_LABEL_COLUMN,
            attach_study_metadata,
            create_study,
            link_dataset_to_study,
        )

        project_id = create_project("Manifest test")
        dataset_id = add_dataset(project_id, "Mica", "mica", "mica.xlsx", "Data", "sha", "", 1)
        replace_dataset_rows(dataset_id, pd.DataFrame([{"Sample": "PG-1", "SiO2": 40.0}]))
        study_id = create_study(
            project_id,
            source_type="article",
            citation="Ivanov et al., 2020",
            doi="10.1000/example",
        )
        link_dataset_to_study(dataset_id, study_id, source_table="Table S1")
        from petrolab.db import load_dataset_dataframe

        dataframe = attach_study_metadata(load_dataset_dataframe(dataset_id, include_meta=True))
        manifest = build_selection_manifest(
            kind="xy_figure",
            dataframe=dataframe,
            dataset_ids=[dataset_id],
            filters={"Sample": ["PG-1"], "visible_sources": ["Ivanov et al., 2020"], "hidden_sources": []},
            recipe={"x": "SiO2", "y": "TiO2"},
        )
        assert manifest["analysis_ids"] == dataframe["_analysis_id"].astype(str).tolist()
        assert manifest["dataset_ids"] == [dataset_id]
        source_fields = {row["field"]: row["values"] for row in manifest["sources"]}
        assert source_fields[SOURCE_LABEL_COLUMN] == ["Ivanov et al., 2020"]
        assert manifest["filters"]["visible_sources"] == ["Ivanov et al., 2020"]
        workbook = workbook_with_manifest({"Points": dataframe}, manifest)
        book = load_workbook(filename=BytesIO(workbook), read_only=True)
        assert {"Points", "Manifest"}.issubset(book.sheetnames)
        assert "petrolab-publication-manifest/v1" in str(book["Manifest"]["A2"].value)

    print("publication manifest tests: OK")


if __name__ == "__main__":
    main()

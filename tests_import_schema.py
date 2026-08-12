from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

_tmp = tempfile.TemporaryDirectory()
os.environ["PETROLAB_DATA_DIR"] = str(Path(_tmp.name) / "petrolab_data")

from petrolab.db import create_project, get_dataset, load_dataset_dataframe, load_unified_analyses
from petrolab.services.import_service import import_linked_sheets, refresh_dataset_from_source

root = Path(_tmp.name)
workbook = root / "heterogeneous.xlsx"
with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
    pd.DataFrame(
        {
            "Sample": ["A1", "A2"],
            "Point": ["1", "2"],
            "Generation": ["core", "rim"],
            "SiO2": [40.0, 41.0],
            "MgO": [20.0, 19.0],
        }
    ).to_excel(writer, sheet_name="Standard", index=False)
    pd.DataFrame(
        {
            "FeOt": [8.0, 9.0],
            "Генерация": ["N-X1", "N-LF"],
            "Point No": ["1", "2"],
            "MgO": [18.0, 17.0],
            "Sample ID": ["B1", "B2"],
            "SiO₂ (wt. %)": [42.0, 43.0],
            "Rb ppb": [1500.0, 2400.0],
            "La ppm": [50.0, 60.0],
        }
    ).to_excel(writer, sheet_name="Legacy", index=False)

project_id = create_project("Heterogeneous import test")
result = import_linked_sheets(
    project_id=project_id,
    path=workbook,
    sheet_names=["Standard", "Legacy"],
    mineral_key="generic",
    dataset_name="Mixed",
    header_row=1,
    semantic_maps={
        "Standard": {"Sample": "Sample", "Point": "Point", "Generation": "Generation"},
        "Legacy": {"Sample": "Sample ID", "Point": "Point No", "Generation": "Генерация"},
    },
)
assert result.count == 2
standard_id, legacy_id = result.dataset_ids

legacy = load_dataset_dataframe(legacy_id, include_meta=True)
assert legacy["Sample"].tolist() == ["B1", "B2"]
assert legacy["Generation"].tolist() == ["N-X1", "N-LF"]
assert legacy["Point"].astype(str).tolist() == ["1", "2"]
assert legacy["SiO2"].tolist() == [42.0, 43.0]
assert legacy["FeOt"].tolist() == [8.0, 9.0]
assert legacy["Rb [µg/g]"].tolist() == [1.5, 2.4]
assert legacy["La [µg/g]"].tolist() == [50.0, 60.0]
legacy_ids = dict(zip(legacy["Sample"], legacy["_analysis_id"]))

metadata = get_dataset(legacy_id)
column_map = json.loads(metadata["column_map_json"])
assert column_map["Generation"]["original"] == "Генерация"
assert column_map["Sample"]["original"] == "Sample ID"
assert column_map["SiO2"]["original"] == "SiO₂ (wt. %)"
assert column_map["FeOt"]["original"] == "FeOt"
assert column_map["Rb [µg/g]"]["source_unit"].lower() == "ppb"
assert column_map["Rb [µg/g]"]["to_canonical_factor"] == 0.001

unified = load_unified_analyses(project_id, [standard_id, legacy_id])
assert "Generation" in unified.columns
assert "Sample" in unified.columns
assert "Rb [µg/g]" in unified.columns
assert set(unified["Sample"].dropna().astype(str)) == {"A1", "A2", "B1", "B2"}

# Reorder physical columns and modify values. IDs must survive.
book = openpyxl.load_workbook(workbook)
legacy_ws = book["Legacy"]
rows = list(legacy_ws.values)
headers = list(rows[0])
data_rows = [list(row) for row in rows[1:]]
new_order = [
    headers.index("Sample ID"), headers.index("SiO₂ (wt. %)"), headers.index("Генерация"),
    headers.index("Point No"), headers.index("Rb ppb"), headers.index("La ppm"),
    headers.index("FeOt"), headers.index("MgO"),
]
legacy_ws.delete_cols(1, legacy_ws.max_column)
for col_idx, source_idx in enumerate(new_order, start=1):
    legacy_ws.cell(row=1, column=col_idx, value=headers[source_idx])
    for row_idx, row in enumerate(data_rows, start=2):
        legacy_ws.cell(row=row_idx, column=col_idx, value=row[source_idx])
legacy_ws.cell(row=2, column=3, value="N-X2")
legacy_ws.cell(row=2, column=2, value=44.5)
book.save(workbook)
book.close()

refresh = refresh_dataset_from_source(legacy_id)
assert refresh.reused_count == 2
assert refresh.new_count == 0
refreshed = load_dataset_dataframe(legacy_id, include_meta=True)
assert dict(zip(refreshed["Sample"], refreshed["_analysis_id"])) == legacy_ids
assert refreshed.loc[refreshed["Sample"] == "B1", "Generation"].iloc[0] == "N-X2"
assert float(refreshed.loc[refreshed["Sample"] == "B1", "SiO2"].iloc[0]) == 44.5
refreshed_map = json.loads(get_dataset(legacy_id)["column_map_json"])
assert refreshed_map["Sample"]["column_index"] == 1
assert refreshed_map["SiO2"]["column_index"] == 2
assert refreshed_map["Generation"]["column_index"] == 3

# Now rename the generation header, sort existing rows, and insert a new analysis at the top.
# Stable semantic keys must preserve B1/B2 IDs despite new source-row numbers.
replacement = pd.DataFrame(
    {
        "Sample ID": ["B3", "B2", "B1"],
        "SiO₂ (wt. %)": [41.2, 43.0, 44.5],
        "Generation": ["new", "N-LF", "N-X2"],
        "Point No": ["3", "2", "1"],
        "Rb ppb": [900.0, 2400.0, 1500.0],
        "La ppm": [45.0, 60.0, 50.0],
        "FeOt": [7.5, 9.0, 8.0],
        "MgO": [19.0, 17.0, 18.0],
    }
)
with pd.ExcelWriter(workbook, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    replacement.to_excel(writer, sheet_name="Legacy", index=False)

moved = refresh_dataset_from_source(legacy_id)
assert moved.row_count == 3
assert moved.reused_count == 2
assert moved.new_count == 1
assert moved.removed_count == 0
assert moved.moved_rows_detected
assert "Generation" in moved.recovered_roles

after_move = load_dataset_dataframe(legacy_id, include_meta=True)
ids_after = dict(zip(after_move["Sample"], after_move["_analysis_id"]))
assert ids_after["B1"] == legacy_ids["B1"]
assert ids_after["B2"] == legacy_ids["B2"]
assert ids_after["B3"] not in set(legacy_ids.values())
assert int(after_move.loc[after_move["Sample"] == "B1", "_source_row"].iloc[0]) == 4
assert float(after_move.loc[after_move["Sample"] == "B3", "Rb [µg/g]"].iloc[0]) == 0.9

print("heterogeneous import schema tests: OK")
_tmp.cleanup()

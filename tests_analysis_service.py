from __future__ import annotations

import os
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd


def _write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, index=False, sheet_name=sheet_name)


def _change(row: pd.Series, column: str, new_value) -> dict:
    return {
        "analysis_id": str(row["_analysis_id"]),
        "dataset_id": int(row["_dataset_id"]),
        "source_row": int(row["_source_row"]),
        "column_name": column,
        "old_value": row.get(column),
        "new_value": new_value,
    }


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="petrolab_analysis_service_") as tmp:
        base = Path(tmp)
        os.environ["PETROLAB_DATA_DIR"] = str(base / "data")

        from petrolab.db import create_project, ensure_storage, list_change_log, load_unified_analyses
        from petrolab.services.analysis_service import save_changes_and_sync
        from petrolab.services.import_service import import_linked_sheets

        ensure_storage()
        project_id = create_project("Analysis service test", "Automated")

        frame_a = pd.DataFrame({
            "Sample": ["A1", "A2"], "SiO2": [40.0, 41.0], "MgO": [20.0, 19.0],
            "FeO": [8.0, 9.0], "Ba ppb": [1200.0, 1300.0],
        })
        frame_b = pd.DataFrame({
            "Sample": ["B1", "B2"], "SiO2": [42.0, 43.0], "MgO": [18.0, 17.0],
            "FeO": [10.0, 11.0],
        })
        file_a = base / "a.xlsx"; file_b = base / "b.xlsx"
        _write_workbook(file_a, {"Mica": frame_a}); _write_workbook(file_b, {"Mica": frame_b})

        ds_a = import_linked_sheets(
            project_id=project_id, path=file_a, sheet_names=["Mica"], mineral_key="mica",
            dataset_name="A", header_row=1,
        ).dataset_ids[0]
        ds_b = import_linked_sheets(
            project_id=project_id, path=file_b, sheet_names=["Mica"], mineral_key="mica",
            dataset_name="B", header_row=1,
        ).dataset_ids[0]

        unified = load_unified_analyses(project_id, [ds_a, ds_b])
        row_a = unified[unified["_dataset_id"] == ds_a].iloc[0]
        row_b = unified[unified["_dataset_id"] == ds_b].iloc[0]
        assert float(row_a["Ba [µg/g]"]) == 1.2

        # Full preflight: a non-source derived column must prevent every workbook write.
        invalid_result = save_changes_and_sync([
            _change(row_a, "SiO2", 55.0), _change(row_b, "Mg#", 0.777),
        ])
        assert not invalid_result.ok
        wb = openpyxl.load_workbook(file_a, data_only=True); assert wb["Mica"]["B2"].value == 40.0; wb.close()
        wb = openpyxl.load_workbook(file_b, data_only=True); assert wb["Mica"]["B2"].value == 42.0; wb.close()
        unchanged = load_unified_analyses(project_id, [ds_a, ds_b])
        assert float(unchanged[unchanged["_dataset_id"] == ds_a].iloc[0]["SiO2"]) == 40.0

        valid_result = save_changes_and_sync([
            _change(row_a, "SiO2", 44.4), _change(row_b, "SiO2", 46.6),
        ])
        assert valid_result.ok, valid_result.errors
        assert valid_result.synced_files == 2 and len(valid_result.backup_paths) == 2
        assert all(Path(path).exists() for path in valid_result.backup_paths)
        wb = openpyxl.load_workbook(file_a, data_only=True); assert wb["Mica"]["B2"].value == 44.4; wb.close()
        wb = openpyxl.load_workbook(file_b, data_only=True); assert wb["Mica"]["B2"].value == 46.6; wb.close()
        updated = load_unified_analyses(project_id, [ds_a, ds_b])
        assert float(updated[updated["_dataset_id"] == ds_a].iloc[0]["SiO2"]) == 44.4
        assert float(updated[updated["_dataset_id"] == ds_b].iloc[0]["SiO2"]) == 46.6

        trace_row = updated[updated["_dataset_id"] == ds_a].iloc[0]
        trace_result = save_changes_and_sync([_change(trace_row, "Ba [µg/g]", 2.5)])
        assert trace_result.ok, trace_result.errors
        wb = openpyxl.load_workbook(file_a, data_only=True)
        headers = [cell.value for cell in wb["Mica"][1]]
        ba_col = headers.index("Ba ppb") + 1
        assert wb["Mica"].cell(row=2, column=ba_col).value == 2500.0
        wb.close()
        trace_updated = load_unified_analyses(project_id, [ds_a])
        assert float(trace_updated.iloc[0]["Ba [µg/g]"]) == 2.5

        logged = list_change_log(limit=20)
        synced_rows = [row for row in logged if int(row["synced_to_source"]) == 1]
        assert len(synced_rows) >= 3 and all(row["source_backup"] for row in synced_rows[:3])

        multi = base / "multi.xlsx"
        _write_workbook(multi, {"Core": frame_a, "Rim": frame_b})
        multi_result = import_linked_sheets(
            project_id=project_id, path=multi, sheet_names=["Core", "Rim"], mineral_key="mica",
            dataset_name="Multi", header_row=1,
        )
        multi_df = load_unified_analyses(project_id, list(multi_result.dataset_ids))
        core = multi_df[multi_df["Лист"] == "Core"].iloc[0]
        rim = multi_df[multi_df["Лист"] == "Rim"].iloc[0]
        one_file_result = save_changes_and_sync([
            _change(core, "SiO2", 48.1), _change(rim, "SiO2", 49.2),
        ])
        assert one_file_result.ok, one_file_result.errors
        assert one_file_result.synced_files == 1 and len(one_file_result.backup_paths) == 1
        wb = openpyxl.load_workbook(multi, data_only=True)
        assert wb["Core"]["B2"].value == 48.1 and wb["Rim"]["B2"].value == 49.2
        wb.close()
        print("analysis service tests: OK")


if __name__ == "__main__":
    main()

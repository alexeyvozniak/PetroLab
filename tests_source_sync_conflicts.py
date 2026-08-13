from __future__ import annotations

import json
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from petrolab.io_utils import sha256_file
from petrolab.sources import validate_sync_change


with tempfile.TemporaryDirectory() as tmp:
    path = Path(tmp) / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Mica"
    worksheet.cell(1, 1, "Sample")
    worksheet.cell(1, 2, "SiO2")
    worksheet.cell(2, 1, "A1")
    worksheet.cell(2, 2, 40.0)
    workbook.save(path)
    workbook.close()

    dataset = {
        "id": 1,
        "name": "Mica",
        "sync_enabled": 1,
        "source_path": str(path),
        "source_sha256": sha256_file(path),
        "source_sheet": "Mica",
        "column_map_json": json.dumps(
            {
                "SiO2": {
                    "column_index": 2,
                    "to_source_factor": 1.0,
                }
            }
        ),
    }
    change = {
        "analysis_id": "a1",
        "source_row": 2,
        "column_name": "SiO2",
        "old_value": 40.0,
        "new_value": 41.0,
    }

    validate_sync_change(dataset, change)

    workbook = load_workbook(path)
    workbook["Mica"].cell(2, 2, 42.0)
    workbook.save(path)
    workbook.close()

    try:
        validate_sync_change(dataset, change)
    except ValueError as exc:
        assert "изменён вне ПетроЛаба" in str(exc)
    else:
        raise AssertionError("External workbook changes must block source sync")

print("source sync conflict tests: OK")
